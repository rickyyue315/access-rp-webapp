"""RP Maintenance 參數轉換服務 (Zeabur)。

上傳 'RP_Maintenance (input)' 檔案 (.txt/.csv/.xls/.xlsx)，
即時套用 Access MDB 的計算邏輯，下載 'RP_Maintenance (output)' (.xlsx)。

啟動:  python app.py   (或由 gunicorn app:app 啟動)
"""
import io
import os
from datetime import datetime

from flask import (Flask, render_template, request, send_file,
                   jsonify, flash, redirect, url_for)

import input_parser
import transform
from transform import OUTPUT_COLUMNS

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024  # 32 MB
app.secret_key = os.environ.get("SECRET_KEY", "rp-maintenance-zeabur")

ALLOWED_EXT = {".txt", ".csv", ".xls", ".xlsx"}


def _allowed(filename):
    return os.path.splitext(filename)[1].lower() in ALLOWED_EXT


@app.route("/")
def index():
    return render_template("index.html",
                           sites=len(transform.WAREHOUSE_CALENDAR),
                           mss=len(transform.MSS_LIST),
                           moq=len(transform.D001_MOQ))


@app.route("/process", methods=["POST"])
def process():
    file = request.files.get("file")
    if not file or not file.filename:
        flash("請選擇一個檔案。", "error")
        return redirect(url_for("index"))
    if not _allowed(file.filename):
        flash(f"不支援的格式，僅接受 {', '.join(sorted(ALLOWED_EXT))}。", "error")
        return redirect(url_for("index"))

    # 處理 (選用) ABC 覆寫檔
    abc_file = request.files.get("abc_file")
    if abc_file and abc_file.filename:
        abc_bytes = abc_file.read()
        try:
            abc_header, abc_rows = input_parser.parse_file(abc_file.filename, abc_bytes)
            transform.set_abc_override(abc_rows)
        except Exception as e:
            flash(f"ABC 覆寫檔讀取失敗 (已略過): {e}", "warn")

    raw = file.read()
    try:
        header, rows = input_parser.parse_file(file.filename, raw)
    except Exception as e:
        flash(f"檔案剖析失敗: {e}", "error")
        return redirect(url_for("index"))

    if not header:
        flash("檔案中找不到表頭，請確認檔案格式。", "error")
        return redirect(url_for("index"))

    output_rows, stats = transform.transform_rows(rows, header)
    data = transform.rows_to_lists(output_rows)

    # 產出 xlsx
    xlsx_bytes = _build_xlsx(data)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    download_name = f"RP_Maintenance (output)_{stamp}.xlsx"
    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )


@app.route("/api/process", methods=["POST"])
def api_process():
    """JSON / multipart API，回傳處理統計 (不含檔案，供程式化呼叫)。"""
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "no file"}), 400
    if not _allowed(file.filename):
        return jsonify({"ok": False, "error": "unsupported format"}), 400
    header, rows = input_parser.parse_file(file.filename, file.read())
    output_rows, stats = transform.transform_rows(rows, header)
    # 回傳前幾筆預覽 + 統計
    preview = output_rows[:5]
    return jsonify({
        "ok": True,
        "stats": {k: v for k, v in stats.items()},
        "columns": OUTPUT_COLUMNS,
        "preview": preview,
    })


def _build_xlsx(data):
    """data: [header_list, row1, row2, ...] -> xlsx bytes。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Final Result"

    header = data[0]
    ws.append(header)
    for row in data[1:]:
        ws.append(row)

    # 表頭樣式
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="305496")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # 資料列樣式 + 欄寬
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(header) + 1):
            ws.cell(row=r, column=c).border = border

    widths = {
        "Site": 8, "Article": 16, "Article Description": 36,
        "Brand": 8, "MC": 10, "MC Description": 22,
        "New Planning Cycle": 16, "New Delivery Cycle": 16,
        "New Traget Coverage": 18, "New ABC Indicator": 16,
        "A QTY": 8, "B QTY": 8, "C QTY": 8,
    }
    for c, name in enumerate(header, start=1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = \
            widths.get(name, 14)
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)
