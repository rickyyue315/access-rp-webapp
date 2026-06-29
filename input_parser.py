"""讀取 RP_Maintenance (input) 檔案。

支援格式：
  - .txt  (Tab 分隔文字，最常見、與 SAP 匯出一致)
  - .csv  (逗號或分號分隔)
  - .xls / .xlsx (Excel)

回傳 (header, rows)，header 為字串清單，rows 為二維字串清單 (不含表頭)。
"""
import csv
import io


def _clean(val):
    """去除前後空白，並移除被雙引號整個包住的引號（SAP 匯出常見）。

    例如 '"EDT, N/A, 50ML   "' -> 'EDT, N/A, 50ML'
    """
    if val is None:
        return ""
    s = str(val).strip()
    # 若整欄被雙引號包住，移除外層引號（保留內部逗號）
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        s = s[1:-1].strip()
    return s


def parse_file(filename, file_bytes):
    """依副檔名剖析上傳檔案。回傳 (header, rows)。"""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return _parse_excel(file_bytes, name.endswith(".xlsx"))
    # 文字類 (.txt / .csv)
    return _parse_text(file_bytes)


def _parse_text(file_bytes):
    # 嘗試 UTF-8-sig -> utf-8 -> big5 -> latin-1
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp950", "big5", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = file_bytes.decode("latin-1", errors="ignore")

    # 偵測分隔符號：Tab 優先 (SAP .txt)，否則 sniff CSV
    sample = text.split("\n", 1)[0]
    if "\t" in sample:
        delimiter = "\t"
    elif ";" in sample and sample.count(";") > sample.count(","):
        delimiter = ";"
    else:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not all_rows:
        return [], []
    header = [_clean(c) for c in all_rows[0]]
    rows = [[_clean(c) for c in r] for r in all_rows[1:]]
    return header, rows


def _parse_excel(file_bytes, is_xlsx):
    if is_xlsx:
        return _parse_xlsx(file_bytes)
    return _parse_xls(file_bytes)


def _parse_xlsx(file_bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    matrix = []
    for row in ws.iter_rows(values_only=True):
        matrix.append([_clean(v) for v in row])
    wb.close()
    return _matrix_to_header_rows(matrix)


def _parse_xls(file_bytes):
    import xlrd
    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_index(0)
    matrix = []
    for r in range(ws.nrows):
        matrix.append([_clean(ws.cell_value(r, c)) for c in range(ws.ncols)])
    return _matrix_to_header_rows(matrix)


def _matrix_to_header_rows(matrix):
    # 去除全空列
    cleaned = []
    for row in matrix:
        if any((c or "").strip() for c in row):
            cleaned.append([c if c is not None else "" for c in row])
    if not cleaned:
        return [], []
    max_len = max(len(r) for r in cleaned)
    for r in cleaned:
        while len(r) < max_len:
            r.append("")
    header = cleaned[0]
    rows = cleaned[1:]
    return header, rows
